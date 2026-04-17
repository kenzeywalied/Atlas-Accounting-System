import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime


HEADER_FILL  = PatternFill("solid", start_color="1A1A2E")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT   = Font(bold=True, size=14, color="1A1A2E")
SUBHD_FONT   = Font(bold=True, size=11, color="16213E")
TOTAL_FILL   = PatternFill("solid", start_color="E8F4FD")
TOTAL_FONT   = Font(bold=True, size=10, color="1A1A2E")
ALT_FILL     = PatternFill("solid", start_color="F8FAFC")
BORDER_SIDE  = Side(style="thin", color="CBD5E1")
THIN_BORDER  = Border(bottom=BORDER_SIDE)
NUM_FMT      = '#,##0.00'
NUM_FMT_0    = '#,##0'

def _auto_width(ws, min_w=8, max_w=40):
    for col in ws.columns:
        max_len = max_w
        for cell in col:
            try:
                max_len = min(max_w, max(min_w, len(str(cell.value or ""))))
            except:
                pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2


def export_trial_balance(df: pd.DataFrame, title="Trial Balance") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Trial Balance"

    # Title block
    ws.merge_cells("A1:J1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%B %d, %Y %H:%M')}"
    ws["A2"].font = Font(italic=True, color="64748B", size=9)
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["Code", "Account Name", "Type",
               "Beg Dr", "Beg Cr",
               "Period Dr", "Period Cr",
               "End Dr", "End Cr"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(df.itertuples(), 5):
        fill = ALT_FILL if r % 2 == 0 else None
        vals = [row.code, row.name, row.type,
                row.beg_bal_dr, row.beg_bal_cr,
                row.mv_dr, row.mv_cr,
                row.end_bal_dr, row.end_bal_cr]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if fill:
                cell.fill = fill
            if c >= 4:
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal="right")

    # Totals row
    tr = len(df) + 5
    ws.cell(tr, 1, "TOTAL").font = TOTAL_FONT
    ws.cell(tr, 1).fill = TOTAL_FILL
    for c in range(1, 10):
        ws.cell(tr, c).fill = TOTAL_FILL
    for c, col in [(4,"beg_bal_dr"),(5,"beg_bal_cr"),(6,"mv_dr"),(7,"mv_cr"),(8,"end_bal_dr"),(9,"end_bal_cr")]:
        cell = ws.cell(tr, c, df[col].sum())
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.number_format = NUM_FMT
        cell.alignment = Alignment(horizontal="right")

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_journal_entries(entries_df, lines_fn) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Journal Entries"

    ws.merge_cells("A1:G1")
    ws["A1"] = "Journal Entries Register"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Entry#", "Date", "Description", "Reference", "Status", "Debit", "Credit"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    r = 4
    for _, entry in entries_df.iterrows():
        fill = ALT_FILL if r % 2 == 0 else None
        row_vals = [entry.entry_number, entry.entry_date,
                    entry.description, entry.get("reference",""),
                    entry.status, entry.total_debit, entry.total_credit]
        for c, v in enumerate(row_vals, 1):
            cell = ws.cell(r, c, v)
            if fill:
                cell.fill = fill
            if c >= 6:
                cell.number_format = NUM_FMT
        r += 1

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_income_statement(revenue_df, expense_df, start, end) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Income Statement"

    ws.merge_cells("A1:C1")
    ws["A1"] = "Income Statement"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:C2")
    ws["A2"] = f"For the period {start} to {end}"
    ws["A2"].font = Font(italic=True, color="64748B", size=9)
    ws["A2"].alignment = Alignment(horizontal="center")

    def write_section(title, df, start_row):
        ws.cell(start_row, 1, title).font = SUBHD_FONT
        ws.cell(start_row, 1).fill = PatternFill("solid", start_color="DBEAFE")
        r = start_row + 1
        total = 0
        for _, row in df.iterrows():
            ws.cell(r, 1, row["code"])
            ws.cell(r, 2, row["name"])
            cell = ws.cell(r, 3, row["net"])
            cell.number_format = NUM_FMT
            cell.alignment = Alignment(horizontal="right")
            total += row["net"]
            r += 1
        ws.cell(r, 2, f"Total {title}").font = TOTAL_FONT
        ws.cell(r, 2).fill = TOTAL_FILL
        cell = ws.cell(r, 3, total)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.number_format = NUM_FMT
        cell.alignment = Alignment(horizontal="right")
        return r + 2, total

    rev_end, total_rev = write_section("Revenue", revenue_df, 4)
    exp_end, total_exp = write_section("Expenses", expense_df, rev_end)

    net_income = total_rev - total_exp
    ws.cell(exp_end, 2, "NET INCOME").font = Font(bold=True, size=11, color="FFFFFF")
    ws.cell(exp_end, 2).fill = PatternFill("solid", start_color="1E40AF")
    cell = ws.cell(exp_end, 3, net_income)
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="1E40AF")
    cell.number_format = NUM_FMT
    cell.alignment = Alignment(horizontal="right")

    for col in ["A","B","C"]:
        ws.column_dimensions[col].width = 35

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_balance_sheet(assets_df, liab_df, equity_df, as_of) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    ws.merge_cells("A1:C1")
    ws["A1"] = "Balance Sheet"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:C2")
    ws["A2"] = f"As of {as_of}"
    ws["A2"].font = Font(italic=True, color="64748B", size=9)
    ws["A2"].alignment = Alignment(horizontal="center")

    def write_section(title, df, start_row, color="DBEAFE"):
        ws.cell(start_row, 1, title).font = SUBHD_FONT
        ws.cell(start_row, 1).fill = PatternFill("solid", start_color=color)
        r = start_row + 1
        total = 0
        for _, row in df.iterrows():
            ws.cell(r, 1, row["code"])
            ws.cell(r, 2, row["name"])
            cell = ws.cell(r, 3, row["balance"])
            cell.number_format = NUM_FMT
            cell.alignment = Alignment(horizontal="right")
            total += row["balance"]
            r += 1
        ws.cell(r, 2, f"Total {title}").font = TOTAL_FONT
        ws.cell(r, 2).fill = TOTAL_FILL
        cell = ws.cell(r, 3, total)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.number_format = NUM_FMT
        cell.alignment = Alignment(horizontal="right")
        return r + 2, total

    r, total_assets = write_section("Assets", assets_df, 4, "D1FAE5")
    r, total_liab   = write_section("Liabilities", liab_df, r, "FEE2E2")
    r, total_equity = write_section("Equity", equity_df, r, "FEF3C7")

    for col in ["A","B","C"]:
        ws.column_dimensions[col].width = 35

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()